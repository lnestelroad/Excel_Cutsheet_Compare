import sys
sys.path.insert(0, "C:\\Users\\P2822177\\Desktop\\Modules")

import os
import csv

import pandas as pd
import openpyxl
from openpyxl import load_workbook

#######################################################################################################################
def Export():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    pwd = os.getcwd()
    filenames = os.listdir("%s" % pwd)
    Excel_files = []
    counter = 0

    for element in filenames:
        if '.xlsx' in element and '.xlsx.csv' not in element:
            Excel_files.append(element)

    files = []
    for x in Excel_files:
        wb = openpyxl.load_workbook(x)
        sh = wb.get_active_sheet()
        with open('test-' + str(counter) + '.xlsx.csv', 'w') as f:  # open('test.csv', 'w', newline="") for python 3
            c = csv.writer(f)
            for r in sh.rows:
                c.writerow([cell.value for cell in r])
        f.close()
        files.append('test-' + str(counter) + '.xlsx.csv')
        counter += 1

    print(files)
    return files
#######################################################################################################################
def Del():
    pwd = os.path.dirname(os.path.abspath(__file__))
    filenames = os.listdir("%s" % pwd)

    for element in filenames:
        if '.xlsx.csv' in element or '.csv' in element:
            os.system("rm %s" % element)
#######################################################################################################################
def Import(masterfile):
    pwd = os.path.dirname(os.path.abspath(__file__))
    df = pd.read_csv("{}\\{}".format(pwd, masterfile), sep='\t', index_col=False)

    book = load_workbook('{}\\Master_Cutsheet_Template.xlsx'.format(pwd))
    pwd = pwd.strip().split("\\")

    sheet_name = 'Master'
    with pd.ExcelWriter('{}_Master_Liam.xlsx'.format(pwd[-1]), engine='openpyxl') as writer:

        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        df.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=0, engine='openpyxl')
#######################################################################################################################
def Compare(format):
    masterfile = "Main.csv"
    pwd = os.path.dirname(os.path.abspath(__file__))  #Sets the current directory of the file to a variable for later use
    files = Export()  #brings in all of the text files exported by the Export fucntion
    connections = {}  #creates a blank dictionary to store of the the connection pairs
    source_dupe = {}  #another dictionary is made to store the file in which the connections are documented
    dupe_counter = 0  #initiates a counter to keep track of how many duplicates occure in each run
    main = open('{}\\{}'.format(pwd, masterfile), "w")  #opens the master file in write mode
    # main.write("Type\tHostname\tDevice Type\tCabinet Location\tRack Unit\tSlot Number\tPort Number\tHostname\tDecvice Type\tCabinet Location\tRack Unit\tSlot Number\tPort Number\tCable Type\tCable Length\tCable Color\tInterface Type\n")

    for txt in files:  #iteraters through each text file in the directory
        line_counter = 0  #at the beginning of each file, the line counter is reset to one
        with open(pwd + "\\" +'{}'.format(txt)) as f:  #opens each text file in the current directory
            for line in f:  #iteraters through every line in the current text file
                t = line.strip().split(",")  #splits each element in the line into its own slot in an array
                if t[0] == "Install" or t[0] == "Connected":
                    if format == "Liam":
                        key = (t[1] + "\t" +  t[2] + "\t" + t[6] + "\t" + t[7])
                        value = (t[8] + "\t" +  t[9] + "\t" + t[12] + "\t" + t[13])

                        try:
                            new = (t[0] + "\t" + t[1] + "\t" + t[2] + "\t" + t[4] + "\t" + t[5] + "\t" + t[6] + "\t" + t[7]
                                 + "\t" + t[8] + "\t" + t[9] + "\t" + t[10] + "\t" + t[11] + "\t" + t[12] + "\t" + t[13] + "\t" + t[3] + "\t\t\t" + t[14] + "\n")
                        except IndexError:
                            new = (t[0] + "\t" + t[1] + "\t" + t[2] + "\t" + t[4] + "\t" + t[5] + "\t" + t[6] + "\t" + t[7]
                                 + "\t" + t[8] + "\t" + t[9] + "\t" + t[10] + "\t" + t[11] + "\t" + t[12] + "\t" + t[13] + "\t" + t[3] + "\t\t\t" + "\n")

                    else:
                        key = (t[1] + "\t" +  t[2] + "\t" + t[5] + "\t" + t[6])
                        value = (t[7] + "\t" +  t[8] + "\t" + t[11] + "\t" + t[12])

                        new = (t[0] + "\t" + t[1] + "\t" + t[2] + "\t" + t[3] + "\t" + t[4] + "\t" + t[5] + "\t" + t[6] + "\t" + t[7]
                             + "\t" + t[8] + "\t" + t[9] + "\t" + t[10] + "\t" + t[11] + "\t" + t[12] + "\t" + t[13] + "\t" + t[14] + "\t\t" + t[16] + "\n")

                    if key in connections and value == connections[key]:  #determines if a connection already exists in the dictionary
                        #print("Duplicate detected on line " + str(line_counter) + " in file " + txt + " : source file is " +  source_dupe[connections[key]] + "\n")
                        dupe_counter += 1  #adds one to the duplicate counter
                    elif value in connections and key == connections[value]:  #determines if a mirror already exists in the dictionary
                        #print("Duplicate detected on line " + str(line_counter) + " in file " + txt + " : source file is " +  source_dupe[connections[key]] + "\n")
                        dupe_counter += 1  #adds one to the duplicate counter
                    else:  #if the connection does not already exits...
                        file_line = txt + " : " + str(line_counter)  #saves the connections initial line number and file name
                        connections[key] = value  #the new connection in saved as a new entry in the dictionary
                        source_dupe[connections[key]] =  file_line  #the new connections entry also has in corrisponding file name and line number associated with it

                        main.write(new)  #writes

                    line_counter += 1  #the line counter is increaded by one after every iteration
            f.close()
    main.close()
    print (dupe_counter)
    Import(masterfile)
    # Del()
#######################################################################################################################
Compare("Liam")
